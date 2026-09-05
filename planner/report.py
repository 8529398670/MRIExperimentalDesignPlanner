"""XLSX report generation for the MRI Experimental Design Planner.

The workbook is the deliverable a human planner hands to a scanner technologist
or drops into an IRB / grant appendix: the solved design at every level of the
hierarchy - trial, run, session, experiment, study - the session timelines, the
efficiency diagnostics and the verbatim acquisition cards the design was solved
against.
"""

from __future__ import annotations

import io
from typing import Any, Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .protocols import meta_of, sections_of

# ------------------------------------------------------------- WSU palette
GREEN = "046A38"
GOLD = "CBA052"
DARK_GREEN = "00482B"
DEEP_GOLD = "AE8643"
YELLOW = "F8E08E"
LEAF = "719949"
BEIGE = "DCD59A"
BLACK = "101820"
FADED_GOLD = "E7E3C6"
OFF_WHITE = "F2F1F0"
WHITE = "FFFFFF"

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=WHITE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=OFF_WHITE)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=WHITE)
BODY_FONT = Font(name="Calibri", size=10, color=BLACK)
BODY_BOLD = Font(name="Calibri", size=10, bold=True, color=BLACK)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color=DARK_GREEN)

TITLE_FILL = PatternFill("solid", fgColor=DARK_GREEN)
SECTION_FILL = PatternFill("solid", fgColor=GREEN)
HEADER_FILL = PatternFill("solid", fgColor=GREEN)
BAND_FILL = PatternFill("solid", fgColor=OFF_WHITE)
ALT_FILL = PatternFill("solid", fgColor=FADED_GOLD)
ACCENT_FILL = PatternFill("solid", fgColor=YELLOW)

THIN = Side(style="thin", color=BEIGE)
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(vertical="top", wrap_text=True)
LEFT = Alignment(vertical="center", horizontal="left")


# --------------------------------------------------------------- utilities


def _num(value: Any, default: float = 0.0) -> float:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return default
    if result != result or result in (float("inf"), float("-inf")):
        return default
    return result


def _fmt(value: float, digits: int = 1) -> str:
    return f"{_num(value):,.{digits}f}"


def _title_block(ws, title: str, subtitle: str, span: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    cell = ws.cell(row=2, column=1, value=subtitle)
    cell.font = SUBTITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16
    return 4


def _section(ws, row: int, text: str, span: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1


def _table(
    ws,
    row: int,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    widths: Sequence[int] | None = None,
    wrap_columns: Sequence[int] = (),
    number_formats: Dict[int, str] | None = None,
) -> int:
    number_formats = number_formats or {}
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = GRID
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    ws.row_dimensions[row].height = 18
    row += 1

    for band, record in enumerate(rows):
        for index, value in enumerate(record, start=1):
            cell = ws.cell(row=row, column=index, value=value)
            cell.font = BODY_FONT
            cell.fill = BAND_FILL if band % 2 == 0 else ALT_FILL
            cell.border = GRID
            cell.alignment = WRAP if index in wrap_columns else LEFT
            if index in number_formats:
                cell.number_format = number_formats[index]
        row += 1

    if widths:
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
    return row + 1


def _kv_rows(ws, row: int, pairs: Sequence[Sequence[Any]], label_width: int = 38,
             value_width: int = 34) -> int:
    for band, (label, value) in enumerate(pairs):
        left = ws.cell(row=row, column=1, value=label)
        left.font = BODY_BOLD
        left.fill = BAND_FILL if band % 2 == 0 else ALT_FILL
        left.border = GRID
        left.alignment = LEFT
        right = ws.cell(row=row, column=2, value=value)
        right.font = BODY_FONT
        right.fill = BAND_FILL if band % 2 == 0 else ALT_FILL
        right.border = GRID
        right.alignment = LEFT
        row += 1
    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = value_width
    return row + 1


def _prose(ws, row: int, text: str, width: int = 132) -> int:
    ws.column_dimensions["A"].width = width
    for line in str(text or "").split("\n"):
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        row += 1
    return row


def _experiments(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    return payload.get("experiments", []) or []


def _runs(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [r for r in (payload.get("runs", []) or []) if not r.get("missing")]


def _sessions(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [s for s in (payload.get("sessions", []) or []) if not s.get("missing")]


# ------------------------------------------------------------------ sheets


def _sheet_summary(wb: Workbook, payload: Dict[str, Any]) -> None:
    meta = payload.get("meta", {})
    totals = payload.get("totals", {})
    budget = payload.get("budget", {})

    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = GREEN
    row = _title_block(
        ws,
        meta.get("studyTitle", "MRI Experimental Design"),
        f"Generated {payload.get('generated', '')}  |  "
        f"{meta.get('institution', '')}  |  {meta.get('designId', '')}",
        span=8,
    )

    row = _section(ws, row, "Study", 8)
    row = _kv_rows(ws, row, [
        ["Investigator", meta.get("investigator", "")],
        ["Institution", meta.get("institution", "")],
        ["Participant", meta.get("participantId", "")],
        ["Design ID", meta.get("designId", "")],
        ["Notes", meta.get("notes", "")],
    ])

    row = _section(ws, row, "Totals", 8)
    row = _kv_rows(ws, row, [
        ["Experiments", len(_experiments(payload))],
        ["Sessions", totals.get("sessions", 0)],
        ["Runs", totals.get("runs", 0)],
        ["Trials", totals.get("trials", 0)],
        ["Primary events", totals.get("units", 0)],
        ["Control trials", totals.get("controlTrials", 0)],
        ["Scanner hours available", f"{_fmt(totals.get('totalScannerHours'))} h"],
        ["Usable after contingency", f"{_fmt(totals.get('usableHours'))} h"],
        ["Committed", f"{_fmt(totals.get('committedHours'))} h"],
        ["Functional acquisition", f"{_fmt(totals.get('functionalHours'))} h"],
        ["Setup, structurals and breaks", f"{_fmt(totals.get('overheadHours'))} h"],
        ["Remaining", f"{_fmt(totals.get('remainingHours'))} h"],
        ["Utilisation", f"{_fmt(totals.get('utilisationPct'))} %"],
        ["Raw functional data", f"{_fmt(totals.get('dataVolumeGb'), 2)} GB"],
        ["Sessions per week", totals.get("sessionsPerWeek", 0)],
        ["Weeks needed", f"{totals.get('weeksNeeded', 0)} of {totals.get('weeksAvailable', 0)}"],
        ["Solve mode", budget.get("solveMode", "")],
        ["Constraint flags", totals.get("warningCount", 0)],
    ])

    row = _section(ws, row, "Experiments", 8)
    rows = []
    for experiment in _experiments(payload):
        derived = experiment.get("derived", {})
        unit = experiment.get("unit", {})
        rows.append([
            experiment.get("name", ""),
            unit.get("plural", "trials"),
            derived.get("sessions", 0),
            derived.get("runs", 0),
            derived.get("trials", 0),
            derived.get("units", 0),
            derived.get("targetUnits", 0),
            f"{_fmt(derived.get('totalHours'), 2)} h",
        ])
    row = _table(
        ws, row,
        ["Experiment", "Unit", "Sessions", "Runs", "Trials", "Collected", "Goal", "Hours"],
        rows,
        widths=[34, 16, 11, 9, 11, 12, 11, 12],
    )

    if payload.get("warnings"):
        row = _section(ws, row, "Constraint report", 8)
        for index, warning in enumerate(payload["warnings"], start=1):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value=f"{index}.  {warning}")
            cell.font = NOTE_FONT
            cell.fill = ACCENT_FILL
            cell.alignment = WRAP
            row += 1


def _sheet_experiments(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Experiments")
    ws.sheet_properties.tabColor = GOLD
    row = _title_block(ws, "Experiments", "Session plans, goals and what each one buys", span=8)

    for experiment in _experiments(payload):
        derived = experiment.get("derived", {})
        unit = experiment.get("unit", {})
        row = _section(ws, row, experiment.get("name", "Experiment"), 8)
        if experiment.get("note"):
            cell = ws.cell(row=row, column=1, value=experiment["note"])
            cell.font = NOTE_FONT
            cell.alignment = WRAP
            row += 2

        row = _table(
            ws, row,
            ["Level", "Composition", "Trials", "Duration"],
            [
                [entry.get("level", ""), entry.get("sequence", ""),
                 entry.get("count", 0), entry.get("duration", "")]
                for entry in experiment.get("table", [])
            ],
            widths=[16, 88, 12, 22],
            wrap_columns=(2,),
        )

        if experiment.get("plan"):
            row = _table(
                ws, row,
                ["Session", "Asked for", "Scheduled", "Runs each", "Minutes each",
                 f"{unit.get('plural', 'trials')} each", "Total minutes"],
                [
                    [entry.get("name", ""), entry.get("requested", 0), entry.get("sessions", 0),
                     entry.get("runsEach", 0), entry.get("minutesEach", 0),
                     entry.get("unitsEach", 0), entry.get("minutes", 0)]
                    for entry in experiment["plan"]
                ],
                widths=[34, 12, 12, 12, 14, 16, 14],
            )

        if experiment.get("runs"):
            row = _table(
                ws, row,
                ["Run design", "Trial design", "Card", "Per session", "Total runs",
                 "Trials", unit.get("plural", "trials")],
                [
                    [entry.get("name", ""), entry.get("trialName", ""),
                     entry.get("protocolLabel", ""), entry.get("perSession", 0),
                     entry.get("totalRuns", 0), entry.get("trials", 0), entry.get("units", 0)]
                    for entry in experiment["runs"]
                ],
                widths=[30, 30, 30, 13, 12, 12, 14],
            )

        row = _kv_rows(ws, row, [
            ["Sessions", derived.get("sessions", 0)],
            ["Expected session duration", f"{derived.get('sessionMeanMinutes', 0)} min"],
            ["Longest session", f"{derived.get('sessionMaxMinutes', 0)} min"],
            [f"{unit.get('plural', 'trials')} per session", derived.get("unitsPerSession", 0)],
            ["Goal", derived.get("targetUnits", 0)],
            ["Progress towards the goal", f"{derived.get('targetProgressPct', 0)} %"],
            ["Share of committed time", f"{derived.get('sharePct', 0)} %"],
            ["Data volume", f"{derived.get('gbTotal', 0)} GB"],
        ])


def _sheet_trials(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Trial Designs")
    ws.sheet_properties.tabColor = LEAF
    row = _title_block(ws, "Trial designs", "What one trial looks like, phase by phase", span=6)

    row = _table(
        ws, row,
        ["Trial design", "Objective", "Phases", "Shortest", "Mean", "Longest",
         "Control share", "Used by"],
        [
            [trial.get("name", ""), trial.get("objectiveLabel", ""),
             len(trial.get("phases", [])), f"{trial.get('timing', {}).get('min', 0)} s",
             f"{trial.get('timing', {}).get('mean', 0)} s",
             f"{trial.get('timing', {}).get('max', 0)} s",
             f"{trial.get('controlPct', 0)} %",
             ", ".join(trial.get("usedBy", []))]
            for trial in payload.get("trials", [])
        ],
        widths=[30, 26, 9, 12, 12, 12, 14, 34],
    )

    for trial in payload.get("trials", []):
        row = _section(ws, row, trial.get("name", "Trial design"), 6)
        conditions = trial.get("conditions", {})
        row = _kv_rows(ws, row, [
            ["Objective", trial.get("objectiveLabel", "")],
            ["Condition A", conditions.get("a", "")],
            ["Condition B", conditions.get("b", "")],
            ["Share on condition A", f"{trial.get('conditionBalance', 50)} %"],
            ["Embedded control share", f"{trial.get('controlPct', 0)} %"],
            ["Residual tolerance", f"{trial.get('separationTolerancePct', 0)} %"],
            ["Sequence", trial.get("sequence", "")],
        ], value_width=90)
        row = _table(
            ws, row,
            ["#", "Phase", "Role", "Min (s)", "Max (s)", "Jitter"],
            [
                [index, phase.get("name", ""), phase.get("role", ""),
                 _num(phase.get("min")), _num(phase.get("max")),
                 "yes" if phase.get("jitter") else "no"]
                for index, phase in enumerate(trial.get("phases", []), start=1)
            ],
            widths=[6, 30, 22, 12, 12, 10],
            number_formats={4: "0.0", 5: "0.0"},
        )


def _sheet_runs(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Run Designs")
    ws.sheet_properties.tabColor = DEEP_GOLD
    row = _title_block(ws, "Run designs", "Trials laid out into blocks, bound to a card", span=8)

    row = _table(
        ws, row,
        ["Run design", "Trial design", "Card", "TR (ms)", "Blocks", "Trials/block",
         "Trials/run", "Shortest", "Expected", "Longest", "Volumes", "Scheduled runs"],
        [
            [run.get("name", ""), run.get("trialName", ""), run.get("protocolLabel", ""),
             _num(run.get("trMs")), run.get("structure", {}).get("blocksPerRun", 0),
             run.get("structure", {}).get("trialsPerBlock", 0),
             run.get("derived", {}).get("trialsPerRun", 0),
             f"{_fmt(_num(run.get('derived', {}).get('runMin')) / 60, 2)} min",
             f"{_fmt(_num(run.get('derived', {}).get('runMean')) / 60, 2)} min",
             f"{_fmt(_num(run.get('derived', {}).get('runMax')) / 60, 2)} min",
             run.get("derived", {}).get("volumesPerRun", 0),
             run.get("derived", {}).get("totalRuns", 0)]
            for run in _runs(payload)
        ],
        widths=[28, 26, 26, 10, 9, 12, 11, 12, 12, 12, 10, 13],
    )

    for run in _runs(payload):
        structure = run.get("structure", {})
        derived = run.get("derived", {})
        decode = run.get("decode", {})
        row = _section(ws, row, run.get("name", "Run design"), 8)
        row = _kv_rows(ws, row, [
            ["Trial design", run.get("trialName", "")],
            ["Acquisition card", f"{run.get('protocolLabel', '')}  ({run.get('protocol', '')})"],
            ["TR / TE", f"{_fmt(run.get('trMs'), 0)} / {_fmt(run.get('teMs'), 1)} ms"],
            ["Trials per block", structure.get("trialsPerBlock", 0)],
            ["Blocks per run", structure.get("blocksPerRun", 0)],
            ["Inter-trial gap", f"{structure.get('interTrialGap', 0)} s"],
            ["Inter-block rest", f"{structure.get('interBlockRest', 0)} s"],
            ["Dummy volumes", structure.get("dummyVolumes", 0)],
            ["Lead-in / lead-out", f"{structure.get('leadIn', 0)} / {structure.get('leadOut', 0)} s"],
            ["Condition ordering", decode.get("labelOrder", "")],
            ["Same-condition run length", decode.get("labelRunLength", 1)],
            ["Trials per run", derived.get("trialsPerRun", 0)],
            ["Primary events per run", derived.get("unitsPerRun", 0)],
            ["Volumes per run", derived.get("volumesPerRun", 0)],
            ["Seconds per trial", derived.get("secondsPerTrial", 0)],
            ["Used by", ", ".join(run.get("usedBy", []))],
        ], value_width=52)


def _sheet_sessions(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Sessions")
    ws.sheet_properties.tabColor = BEIGE
    row = _title_block(ws, "Session library", "Every block in the order the console runs it",
                       span=8)

    row = _table(
        ws, row,
        ["Session", "Runs", "Setup", "Functional", "Shortest", "Expected", "Longest",
         "Trials", "Data (GB)", "Scheduled", "Used by"],
        [
            [session.get("name", ""), session.get("runs", 0),
             f"{session.get('setupMinutes', 0)} min",
             f"{session.get('functionalMinutes', 0)} min",
             f"{session.get('minMinutes', 0)} min",
             f"{session.get('meanMinutes', 0)} min",
             f"{session.get('maxMinutes', 0)} min",
             session.get("trials", 0), session.get("gb", 0),
             session.get("scheduled", 0), ", ".join(session.get("usedBy", []))]
            for session in _sessions(payload)
        ],
        widths=[30, 8, 12, 13, 12, 12, 12, 10, 11, 12, 34],
    )

    for session in _sessions(payload):
        row = _section(ws, row, session.get("name", "Session"), 8)
        auto_break = (
            f"{session.get('breakMinutes', 0)} min between adjacent runs"
            if session.get("autoBreak", True)
            else "off - breaks are placed by hand"
        )
        row = _kv_rows(ws, row, [
            ["Setup steps", f"{session.get('overheadMinutes', 0)} min"],
            ["Structural and reference scans", f"{session.get('structuralMinutes', 0)} min"],
            ["Breaks", f"{session.get('breakTotalMinutes', 0)} min"],
            ["Automatic break", auto_break],
            ["Functional acquisition", f"{session.get('functionalMinutes', 0)} min"],
            ["Expected session duration", f"{session.get('meanMinutes', 0)} min"],
            ["Primary events per session", session.get("units", 0)],
        ])
        if session.get("setup"):
            row = _table(
                ws, row,
                ["Setup step", "On", "Minutes"],
                [
                    [entry.get("label", ""), "yes" if entry.get("enabled") else "no",
                     round(_num(entry.get("minutes")), 2)]
                    for entry in session["setup"]
                ],
                widths=[44, 8, 12],
            )
        if session.get("structurals"):
            row = _table(
                ws, row,
                ["Structural / reference card", "On", "Count", "Minutes each", "Minutes"],
                [
                    [entry.get("protocolLabel", ""), "yes" if entry.get("enabled") else "no",
                     entry.get("count", 0), round(_num(entry.get("minutesEach")), 2),
                     round(_num(entry.get("minutes")), 2)]
                    for entry in session["structurals"]
                ],
                widths=[36, 8, 10, 14, 12],
            )
        if session.get("items"):
            row = _table(
                ws, row,
                ["Run design", "Card", "Count", "Minutes each", "Minutes", "Trials"],
                [
                    [entry.get("runName", ""), entry.get("protocolLabel", ""),
                     entry.get("count", 0), entry.get("minutesEach", 0),
                     entry.get("minutes", 0), entry.get("trials", 0)]
                    for entry in session["items"]
                ],
                widths=[30, 30, 10, 14, 12, 10],
            )


def _sheet_timelines(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Session Timelines")
    ws.sheet_properties.tabColor = FADED_GOLD
    row = _title_block(ws, "Session timelines", "Every session, start to finish", span=6)

    for session in _sessions(payload):
        if not session.get("timeline"):
            continue
        row = _section(
            ws, row,
            f"{session.get('name', 'Session')}  -  "
            f"{session.get('meanMinutes', 0)} min expected, "
            f"{session.get('maxMinutes', 0)} min longest",
            6,
        )
        row = _table(
            ws, row,
            ["#", "Item", "Card", "Minutes", "Cumulative", "Category"],
            [
                [entry.get("order", 0), entry.get("item", ""),
                 entry.get("protocolLabel", ""), entry.get("minutes", 0),
                 entry.get("cumulative", 0), entry.get("category", "")]
                for entry in session["timeline"]
            ],
            widths=[6, 46, 30, 11, 13, 22],
            number_formats={4: "0.00", 5: "0.00"},
        )


def _sheet_budget(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Budget")
    ws.sheet_properties.tabColor = DEEP_GOLD
    totals = payload.get("totals", {})
    budget = payload.get("budget", {})
    caps = payload.get("caps", {})
    row = _title_block(ws, "Budget and allocation", "Where the scanner time goes", span=7)

    row = _section(ws, row, "Envelope", 7)
    row = _kv_rows(ws, row, [
        ["Total scanner hours", _num(budget.get("totalScannerHours"))],
        ["Contingency reserve", f"{_num(budget.get('contingencyPct'))} %"],
        ["Usable hours", _num(totals.get("usableHours"))],
        ["Committed hours", _num(totals.get("committedHours"))],
        ["Remaining hours", _num(totals.get("remainingHours"))],
        ["Utilisation", f"{_num(totals.get('utilisationPct'))} %"],
        ["Solve mode", budget.get("solveMode", "")],
        ["Overhead charged to the budget",
         "yes" if budget.get("countOverheadAgainstBudget", True) else "no"],
        ["Auto-clamp", "yes" if budget.get("autoClamp") else "no"],
        ["Sessions per week", _num(budget.get("sessionsPerWeek"))],
        ["Weeks available", _num(budget.get("weeksAvailable"))],
    ])

    row = _section(ws, row, "Allocation", 7)
    rows = []
    for experiment in _experiments(payload):
        derived = experiment.get("derived", {})
        rows.append([
            experiment.get("name", ""),
            f"{_fmt(experiment.get('requestedPct'))} %",
            f"{_fmt(derived.get('sharePct'))} %",
            derived.get("sessions", 0),
            _num(derived.get("functionalHours")),
            _num(derived.get("overheadHours")),
            _num(derived.get("totalHours")),
        ])
    rows.append([
        "Total", "100 %", "100 %", totals.get("sessions", 0),
        _num(totals.get("functionalHours")), _num(totals.get("overheadHours")),
        _num(totals.get("committedHours")),
    ])
    row = _table(
        ws, row,
        ["Experiment", "Requested", "Solved share", "Sessions",
         "Functional h", "Overhead h", "Total h"],
        rows,
        widths=[34, 13, 14, 11, 14, 13, 12],
        number_formats={5: "0.00", 6: "0.00", 7: "0.00"},
    )

    row = _section(ws, row, "Constraint envelope", 7)
    row = _kv_rows(ws, row, [
        ["Caps applied to", caps.get("applyTo", "expected")],
        ["Maximum run duration", f"{_num(caps.get('maxRunMinutes'))} min"],
        ["Maximum session duration", f"{_num(caps.get('maxSessionMinutes'))} min"],
        ["Maximum runs per session", _num(caps.get("maxRunsPerSession"))],
        ["Maximum sessions in total", _num(caps.get("maxSessionsTotal"))],
        ["Continuous-scanning limit", f"{_num(caps.get('maxContinuousMinutes'))} min"],
        ["Minimum per experiment", _num(caps.get("minUnitsPerExperiment"))],
    ])


def _sheet_efficiency(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Efficiency")
    ws.sheet_properties.tabColor = YELLOW
    row = _title_block(ws, "Design efficiency", "Simulated at the bound TR, per run design",
                       span=10)

    rows = []
    for run in _runs(payload):
        e = run.get("efficiency") or {}
        rows.append([
            run.get("name", ""),
            f"{_fmt(e.get('sustainPct'))} %",
            f"{_fmt(e.get('saturationIndex'), 2)} x",
            round(_num(e.get("singleTrialEff")), 4),
            f"{_fmt(e.get('carryoverPct'))} %",
            f"{_fmt(e.get('stimulusBleedPct'))} %",
            round(_num(e.get("effAvsB")), 4),
            round(_num(e.get("effResponseVsBaseline")), 4),
            round(_num(e.get("corrStimulusResponse")), 4),
            round(_num(e.get("maxVif")), 3),
            e.get("volumes", 0),
        ])
    row = _table(
        ws, row,
        ["Run design", "Duty cycle", "Stacking", "Single-trial eff.", "Carryover",
         "Stimulus bleed", "A vs B", "Response vs base", "Stim/resp r", "Max VIF",
         "Volumes"],
        rows,
        widths=[28, 12, 11, 16, 12, 14, 11, 16, 12, 10, 10],
    )

    hrf = payload.get("hrf", {})
    row = _section(ws, row, "Timing model", 10)
    row = _kv_rows(ws, row, [
        ["Peak delay", f"{_num(hrf.get('peakDelay'), 6)} s"],
        ["Peak dispersion", _num(hrf.get("peakDispersion"), 1)],
        ["Undershoot delay", f"{_num(hrf.get('undershootDelay'), 16)} s"],
        ["Undershoot dispersion", _num(hrf.get("undershootDispersion"), 1)],
        ["Peak to undershoot ratio", _num(hrf.get("undershootRatio"), 6)],
        ["Response evaluated over", f"{_num(hrf.get('spanSeconds'), 40)} s"],
        ["Residuals read after onset", f"{_num(hrf.get('readLagSeconds'), 5)} s"],
    ])

    objectives = (hrf.get("objectives") or {})
    if objectives:
        row = _table(
            ws, row,
            ["Objective", "Name in the planner", "Residual tolerance", "Pinned recovery"],
            [
                [key, value.get("label", key), f"{_num(value.get('tolerancePct'))} %",
                 (f"{_num(value.get('separationSeconds'))} s"
                  if _num(value.get("separationSeconds")) > 0 else "solved from the HRF")]
                for key, value in objectives.items()
            ],
            widths=[18, 32, 18, 24],
        )

    row = _section(ws, row, "Reading the numbers", 10)
    for note in [
        "Duty cycle - median predicted task signal as a percentage of its 95th percentile. "
        "High means the response never settles, which is what detection wants; near zero "
        "means full recovery, which is what separation wants.",
        "Stacking gain - peak predicted signal divided by the peak of one isolated trial.",
        "Single-trial efficiency - reciprocal mean variance of least-squares-all trial betas.",
        "Carryover - previous response still present at the next stimulus onset.",
        "Stimulus bleed - stimulus response still present inside the response window.",
        "Max VIF - variance inflation across the stimulus and the two condition regressors.",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=note)
        cell.font = NOTE_FONT
        cell.alignment = WRAP
        ws.row_dimensions[row].height = 26
        row += 1


def _sheet_data_volume(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Data Volume")
    ws.sheet_properties.tabColor = BEIGE
    totals = payload.get("totals", {})
    row = _title_block(ws, "Data volume", "Raw functional data at 16 bits per voxel", span=7)

    rows = []
    for run in _runs(payload):
        volume = run.get("dataVolume", {})
        rows.append([
            run.get("name", ""), volume.get("matrix", ""), volume.get("slices", 0),
            volume.get("voxel", ""), volume.get("volumesPerRun", 0),
            _num(volume.get("mbPerRun")), run.get("derived", {}).get("totalRuns", 0),
            _num(volume.get("gbTotal")),
        ])
    rows.append([
        "Total", "", "", "", "", "", "",
        _num(totals.get("dataVolumeGb")),
    ])
    row = _table(
        ws, row,
        ["Run design", "Matrix", "Slices", "Voxel (mm)", "Volumes/run", "MB/run",
         "Runs", "Total GB"],
        rows,
        widths=[28, 14, 9, 22, 13, 12, 10, 12],
        number_formats={6: "0.0", 8: "0.000"},
    )

    row = _section(ws, row, "Per session", 7)
    row = _table(
        ws, row,
        ["Session", "Runs", "Data per session (GB)", "Scheduled", "Total (GB)"],
        [
            [session.get("name", ""), session.get("runs", 0), _num(session.get("gb")),
             session.get("scheduled", 0),
             round(_num(session.get("gb")) * _num(session.get("scheduled")), 3)]
            for session in _sessions(payload)
        ],
        widths=[30, 9, 22, 12, 14],
        number_formats={3: "0.000", 5: "0.000"},
    )

    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=7)
    cell = ws.cell(
        row=row, column=1,
        value="Functional acquisition only, before structurals, physiological logs and "
              "derivatives. Reconstructed volumes are counted as matrix x matrix x slices "
              "at 2 bytes per voxel.",
    )
    cell.font = NOTE_FONT
    cell.alignment = WRAP


def _sheet_methods(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Methods Text")
    ws.sheet_properties.tabColor = DARK_GREEN
    row = _title_block(ws, "Methods text", "Generated from the solved design", span=1)
    _prose(ws, row, payload.get("methodsText", ""))


def _sheet_markdown(wb: Workbook, payload: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Markdown Tables")
    ws.sheet_properties.tabColor = GOLD
    row = _title_block(
        ws, "Markdown tables",
        "Every table in the planner as GitHub-flavoured Markdown, one row per line",
        span=1,
    )
    ws.column_dimensions["A"].width = 150
    for name, text in (payload.get("markdownTables") or {}).items():
        cell = ws.cell(row=row, column=1, value=f"## {name}")
        cell.font = BODY_BOLD
        row += 2
        for line in str(text).split("\n"):
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = Font(name="Consolas", size=9, color=BLACK)
            row += 1
        row += 1


INVALID_SHEET_CHARS = r"[]:*?/\\"


def _sheet_name(wb: Workbook, preferred: str, fallback: str) -> str:
    """Excel rejects []:*?/\\ and titles longer than 31 characters."""
    cleaned = "".join(" " if ch in INVALID_SHEET_CHARS else ch
                      for ch in (preferred or fallback))
    cleaned = " ".join(cleaned.split()).strip("'")[:28] or fallback[:28]
    name = cleaned
    for suffix in range(2, 40):
        if name not in wb.sheetnames:
            return name
        name = f"{cleaned[:26]} {suffix}"
    return fallback[:31]


def _sheet_protocol(wb: Workbook, slug: str, data: Dict[str, Any], bound_to: str) -> None:
    meta = meta_of(data, slug)
    ws = wb.create_sheet(_sheet_name(wb, meta["label"], slug))
    ws.sheet_properties.tabColor = GREEN if bound_to else BEIGE
    subtitle = f"{slug}.json  |  {meta['role']}"
    if bound_to:
        subtitle += f"  |  used by {bound_to}"
    row = _title_block(ws, meta["label"], subtitle, span=3)

    if meta.get("note"):
        cell = ws.cell(row=row, column=1, value=meta["note"])
        cell.font = NOTE_FONT
        cell.alignment = WRAP
        row += 2

    for section in sections_of(data):
        row = _section(ws, row, section, 3)
        records = []
        for record in data[section]:
            indent = int(_num(record.get("indent")))
            records.append([
                ("    " * indent) + str(record.get("parameter", "")),
                str(record.get("value", "")),
                indent,
            ])
        row = _table(ws, row, ["Parameter", "Value", "Indent"], records, widths=[44, 34, 9])


# ------------------------------------------------------------------- entry


def build_workbook(payload: Dict[str, Any], protocols: Dict[str, Any]) -> bytes:
    """Render the full planner report and return the .xlsx bytes."""
    wb = Workbook()
    _sheet_summary(wb, payload)
    _sheet_experiments(wb, payload)
    _sheet_trials(wb, payload)
    _sheet_runs(wb, payload)
    _sheet_sessions(wb, payload)
    _sheet_timelines(wb, payload)
    _sheet_budget(wb, payload)
    _sheet_efficiency(wb, payload)
    _sheet_data_volume(wb, payload)
    _sheet_methods(wb, payload)
    _sheet_markdown(wb, payload)

    # Which cards the design actually touches, and what touches them.
    bindings: Dict[str, List[str]] = {}
    for run in _runs(payload):
        if run.get("protocol"):
            bindings.setdefault(run["protocol"], []).append(run.get("name", ""))
    for session in _sessions(payload):
        for entry in session.get("structurals", []):
            if entry.get("enabled") and entry.get("protocol"):
                bindings.setdefault(entry["protocol"], []).append(session.get("name", ""))

    for slug, data in (protocols or {}).items():
        if not isinstance(data, dict) or "_error" in data:
            continue
        bound = ", ".join(dict.fromkeys(bindings.get(slug, [])))
        _sheet_protocol(wb, slug, data, bound)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A3"

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
